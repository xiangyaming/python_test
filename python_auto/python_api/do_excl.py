# -*- coding:utf-8 -*-
# @FileName  :do_excl.py
# @Time      :2020/6/7 0007 01:15
# @Author    :xiaoming
# -----------------------------------------------------


from openpyxl   import  load_workbook
#只支持.xlsx格式的文件

#  新建EXCEL只能在桌面新建复制过来，或者通过目录新建，不能NEW--pythonfile
#   1,打开Excel
wb=load_workbook('test.xlsx')   #Open the given filename and return the workbook

#   2，定位表单
sheet=wb['python']  #定位表单，返回一个表单对象

#   3，定位单元格  行-列
res=sheet.cell(1,1).value

print('拿到的结果是：',res)
print('最大行数是：',sheet.max_row)   #最大的行数
print('最大列数是：',sheet.max_column)    #最大的列数

#Excel里面的数据类型，数字还是数字，其他都为字符串
print(type(res))